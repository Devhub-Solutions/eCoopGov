"""
Excel Import Service — Thiết kế cho file thực tế Trường Phát

Cấu trúc file:
  - Dòng 1-3: Title/subtitle (bỏ qua)
  - Dòng 4-5: Header merge 2 dòng
  - Dòng 6+:  Data

Giải quyết header merge:
  - Đọc cả dòng 4 lẫn dòng 5
  - Ưu tiên dòng 5 (sub-header cụ thể hơn)
  - Fallback dòng 4 nếu dòng 5 trống
  - Thử ghép "dòng4 dòng5" nếu cả 2 đều cần
"""
import asyncio
import json
import uuid
from pathlib import Path
import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.logging import logger

CHUNK_SIZE = 500

# ─── Mapping chính xác theo file thực tế ─────────────────────────────────────

PHUONG_TIEN_MAP = {
    "tt": None,   # Số thứ tự - bỏ qua

    # Cột B
    "biển kiểm soát": "bien_so",
    "bien kiem soat": "bien_so",

    # Cột C
    "năm sản xuất": "nam_san_xuat",
    "nam san xuat": "nam_san_xuat",

    # Cột D
    "sức chứa (người)": "so_cho",
    "suc chua (nguoi)": "so_cho",
    "sức chứa": "so_cho",
    "suc chua": "so_cho",

    # Cột E
    "màu sơn": "mau_xe",
    "mau son": "mau_xe",

    # Cột F
    "nhãn hiệu xe": "hang_xe",
    "nhan hieu xe": "hang_xe",
    "nhãn hiệu": "hang_xe",

    # Cột G
    "loại hình hoạt động": "loai_hinh_hoat_dong",
    "loai hinh hoat dong": "loai_hinh_hoat_dong",

    # Cột H
    "tuyến khai thác": "tuyen_khai_thac",
    "tuyen khai thac": "tuyen_khai_thac",

    # Cột I
    "ngày hết hạn đăng kiểm": "han_dang_kiem",
    "ngay het han dang kiem": "han_dang_kiem",

    # Cột J
    "ngày hết hạn phù hiệu": "han_phu_hieu",
    "ngay het han phu hieu": "han_phu_hieu",

    # Cột K - BHXH TNDS
    "ngày hết hạn bảo hiểm tnds": "han_bao_hiem",
    "ngay het han bao hiem tnds": "han_bao_hiem",
    "bảo hiểm tnds": "han_bao_hiem",
    "hạn bảo hiểm tnds": "han_bao_hiem",

    # Thiết bị GSHT - sub-headers dòng 5
    "tên thiết bị": "gsht_ten",
    "ten thiet bi": "gsht_ten",
    "đơn vị cung cấp thiết bị": "gsht_don_vi",
    "don vi cung cap thiet bi": "gsht_don_vi",
    "địa chỉ truy cập": "gsht_dia_chi",
    "dia chi truy cap": "gsht_dia_chi",
    "mật khẩu truy cập": "gsht_mat_khau",
    "mat khau truy cap": "gsht_mat_khau",

    # Loại phương tiện - sub-headers dòng 5
    "thuộc sở hữu của đơn vị": "loai_so_huu",
    "thuoc so huu cua don vi": "loai_so_huu",
    "phương tiện đi thuê": "loai_di_thue",
    "phuong tien di thue": "loai_di_thue",

    # Ghi chú
    "ghi chú": "ghi_chu",
    "ghi chu": "ghi_chu",
}

LAI_XE_MAP = {
    "tt": None,

    # Cột B
    "họ và tên": "ho_ten",
    "ho va ten": "ho_ten",

    # Nhiệm vụ - sub-headers dòng 5
    "lái xe": "nhiem_vu_lai_xe",
    "lai xe": "nhiem_vu_lai_xe",
    "nhân viên phục vụ trên xe": "nhiem_vu_nv_phuc_vu",
    "nhan vien phuc vu tren xe": "nhiem_vu_nv_phuc_vu",

    # GPLX - sub-headers dòng 5
    "hạng": "hang_gplx",
    "hang": "hang_gplx",
    "ngày hết hạn": "han_gplx",
    "ngay het han": "han_gplx",

    # Hợp đồng lao động - sub-headers dòng 5
    "ngày ký": "hop_dong_ngay_ky",
    "ngay ky": "hop_dong_ngay_ky",
    "loại hợp đồng": "hop_dong_loai",
    "loai hop dong": "hop_dong_loai",

    # Cột I
    "đóng bhxh, bhyt": "dong_bhxh_bhyt",
    "dong bhxh, bhyt": "dong_bhxh_bhyt",
    "đóng bhxh bhyt": "dong_bhxh_bhyt",
    "có": "dong_bhxh_bhyt",  # giá trị Có/Không

    # Khám sức khỏe - sub-headers dòng 5
    "ngày khám": "ksk_ngay_kham",
    "ngay kham": "ksk_ngay_kham",
    "kết quả": "ksk_ket_qua",
    "ket qua": "ksk_ket_qua",

    # Tập huấn - sub-headers dòng 5
    "ngày tập huấn": "tap_huan_ngay",
    "ngay tap huan": "tap_huan_ngay",
    "đơn vị tập huấn": "tap_huan_don_vi",
    "don vi tap huan": "tap_huan_don_vi",
    "số gcn tập huấn": "tap_huan_so_gcn",
    "so gcn tap huan": "tap_huan_so_gcn",
    "số gcn": "tap_huan_so_gcn",
    "so gcn": "tap_huan_so_gcn",

    # Ghi chú
    "ghi chú": "ghi_chu",
    "ghi chu": "ghi_chu",
}

REQUIRED_FIELDS = {
    "phuong_tien": ["bien_so"],
    "lai_xe": ["ho_ten"],
}

# Defaults theo đúng file Trường Phát
DEFAULT_HEADER_START_ROW = 4
DEFAULT_DATA_START_ROW   = 6


# ─── Main import function ─────────────────────────────────────────────────────

async def import_excel(
    db: AsyncSession,
    filepath: str | Path,
    target_table: str,
    job_id: str,
    header_row: int = DEFAULT_HEADER_START_ROW,
    data_start_row: int = DEFAULT_DATA_START_ROW,
    preview_mode: bool = False,
    chunk_size: int = CHUNK_SIZE,
) -> dict:
    filepath = Path(filepath)
    col_map  = PHUONG_TIEN_MAP if target_table == "phuong_tien" else LAI_XE_MAP
    required = REQUIRED_FIELDS[target_table]
    stats    = {"total": 0, "success": 0, "errors": 0, "error_details": []}

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    # ── Bước 1: Parse multi-row header (dòng 4 + dòng 5) ────────────────────
    # raw_headers[col_idx] = ["text_dong4", "text_dong5"]  (có thể thiếu 1 trong 2 nếu merged)
    raw_headers: dict[int, list[str]] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < header_row:
            continue
        if row_idx >= data_start_row:
            break
        for col_idx, val in enumerate(row):
            if val is not None:
                s = str(val).strip()
                if s:
                    raw_headers.setdefault(col_idx, []).append(s)

    # Map col_idx → DB field
    # Thứ tự ưu tiên: sub-header (dòng5) → main-header (dòng4) → ghép → bỏ ngoặc
    headers: dict[int, str] = {}
    unmatched: list[str]    = []

    for col_idx, parts in raw_headers.items():
        candidates = [
            parts[-1],                         # dòng cuối (sub-header) — cụ thể nhất
            parts[0],                          # dòng đầu (main header)
            " ".join(parts),                   # ghép cả 2
            parts[-1].split("(")[0].strip(),   # bỏ nội dung trong ngoặc
            parts[0].split("(")[0].strip(),
        ]
        matched = False
        for cand in candidates:
            key = cand.lower().strip()
            if key in col_map:
                if col_map[key] is not None:   # None = cột TT, bỏ qua
                    headers[col_idx] = col_map[key]
                matched = True
                break
        if not matched:
            unmatched.append(f"col{col_idx}:{parts[-1]!r}")

    if not headers:
        all_texts = {i: parts for i, parts in raw_headers.items()}
        raise ValueError(
            f"Không map được cột nào từ header. "
            f"Headers tìm thấy: {all_texts}. "
            f"Kiểm tra file Excel hoặc báo admin cập nhật mapping."
        )

    if unmatched:
        logger.warning("unmatched_cols", table=target_table, cols=unmatched)
    logger.info("headers_mapped", table=target_table, count=len(headers))

    # ── Bước 2: Stream data từ dòng 6 ───────────────────────────────────────
    chunk: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < data_start_row:
            continue
        if all(v is None for v in row):
            continue   # dòng trống hoàn toàn

        stats["total"] += 1
        record: dict = {"id": str(uuid.uuid4()), "import_job_id": job_id}

        for col_idx, field in headers.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[field] = _clean_value(val)

        errs = _validate(record, required, row_idx)
        if errs:
            stats["errors"] += 1
            stats["error_details"].extend(errs)
            continue

        chunk.append(record)

        if len(chunk) >= chunk_size and not preview_mode:
            stats["success"] += await _bulk_upsert(db, target_table, chunk)
            chunk.clear()
            await asyncio.sleep(0)  # nhường event loop

    # flush cuối
    if chunk and not preview_mode:
        stats["success"] += await _bulk_upsert(db, target_table, chunk)
    elif preview_mode:
        stats["success"] = len(chunk)

    wb.close()
    logger.info("import_done", table=target_table,
                total=stats["total"], success=stats["success"], errors=stats["errors"])
    return stats


# ─── Bulk upsert ─────────────────────────────────────────────────────────────

async def _bulk_upsert(db: AsyncSession, table: str, records: list[dict]) -> int:
    """MySQL INSERT ... ON DUPLICATE KEY UPDATE — upsert toàn batch trong 1 SQL statement."""
    if not records:
        return 0
    fields = list(records[0].keys())
    # ON DUPLICATE KEY UPDATE: update tất cả cột trừ id (primary key)
    update_fields = [f for f in fields if f != "id"]
    update_clause = ", ".join(f"{f} = VALUES({f})" for f in update_fields)
    sql = text(
        f"INSERT INTO {table} ({', '.join(fields)}) "
        f"VALUES ({', '.join(':' + f for f in fields)}) "
        f"ON DUPLICATE KEY UPDATE {update_clause}"
    )
    await db.execute(sql, records)
    await db.flush()
    logger.info("batch_upserted", table=table, rows=len(records))
    return len(records)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _clean_value(val):
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        return s or None
    if isinstance(val, float) and val == int(val):
        return int(val)   # 2020.0 → 2020
    from datetime import datetime, date
    if isinstance(val, (datetime, date)):
        return str(val)[:10]  # YYYY-MM-DD
    return val


def _validate(record: dict, required: list[str], row_num: int) -> list[dict]:
    return [
        {"row": row_num, "field": f, "error": f"Thiếu giá trị bắt buộc: '{f}'"}
        for f in required if not record.get(f)
    ]
