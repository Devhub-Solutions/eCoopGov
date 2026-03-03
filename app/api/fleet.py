"""
Fleet API — Import Excel & Query linh hoạt

Import:
  POST /fleet/phuong-tien/import   - Upload Excel phương tiện
  POST /fleet/lai-xe/import        - Upload Excel lái xe
  GET  /fleet/import-jobs/{id}     - Xem tiến trình import
  GET  /fleet/import-jobs/{id}/errors - Download lỗi import

Query (filter linh hoạt):
  GET  /fleet/phuong-tien          - List + filter phương tiện
  GET  /fleet/phuong-tien/{id}     - Chi tiết phương tiện
  GET  /fleet/lai-xe               - List + filter lái xe
  GET  /fleet/lai-xe/{id}          - Chi tiết lái xe
  GET  /fleet/stats                - Thống kê tổng hợp
"""
import shutil
import uuid
import json
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, BackgroundTasks, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.core.config import get_settings
from app.core.database import get_db
from app.models.fleet_models import PhuongTien, LaiXe, ImportJob, ImportStatus
from app.services.excel_import_service import import_excel

router = APIRouter(prefix="/fleet", tags=["Fleet - Phương tiện & Lái xe"])
settings = get_settings()


# ════════════════════════════════════════════════════════════════════════════
# IMPORT ENDPOINTS
# ════════════════════════════════════════════════════════════════════════════

@router.post("/phuong-tien/import", summary="Import Excel danh sách phương tiện")
async def import_phuong_tien(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    header_row:    int  = Form(default=4, description="Dòng bắt đầu header (file Trường Phát = 4)"),
    data_start_row: int = Form(default=6, description="Dòng bắt đầu data (file Trường Phát = 6)"),
    preview: bool = Form(default=False, description="Preview: parse nhưng không lưu"),
    db: AsyncSession = Depends(get_db),
):
    return await _handle_import(file, "phuong_tien", header_row, data_start_row, preview, background_tasks, db)


@router.post("/lai-xe/import", summary="Import Excel danh sách lái xe")
async def import_lai_xe(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    header_row:    int  = Form(default=4),
    data_start_row: int = Form(default=6),
    preview: bool = Form(default=False),
    db: AsyncSession = Depends(get_db),
):
    return await _handle_import(file, "lai_xe", header_row, data_start_row, preview, background_tasks, db)


async def _handle_import(file, table, header_row, data_start_row, preview, background_tasks, db):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx hoặc .xls")

    job_id = str(uuid.uuid4())
    filepath = settings.UPLOAD_DIR / f"{job_id}_{file.filename}"
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    job = ImportJob(
        id=job_id,
        target_table=table,
        filename=file.filename,
        status=ImportStatus.PENDING,
        preview_mode=preview,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    db.add(job)
    await db.flush()

    if preview:
        result = await _run_import(job_id, str(filepath), table, header_row, data_start_row, preview=True)
        return {
            "mode": "preview",
            "total_rows":  result["total"],
            "valid_rows":  result["success"],
            "error_rows":  result["errors"],
            "sample_errors": result["error_details"][:10],
            "message": "Preview xong. Gọi lại với preview=false để import thật."
        }
    else:
        background_tasks.add_task(_run_import, job_id, str(filepath), table, header_row, data_start_row, False)
        return {
            "job_id": job_id,
            "status": "pending",
            "message": f"Import đang chạy background. Poll GET /fleet/import-jobs/{job_id}",
        }


@router.get("/import-jobs/{job_id}", summary="Xem tiến trình import")
async def get_import_job(job_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(404, "Import job không tồn tại")

    progress = 0
    if job.total_rows and job.total_rows > 0:
        progress = round(job.processed_rows / job.total_rows * 100, 1)

    return {
        "job_id": job.id,
        "table": job.target_table,
        "filename": job.filename,
        "status": job.status,
        "progress_percent": progress,
        "total_rows": job.total_rows,
        "success_rows": job.success_rows,
        "error_rows": job.error_rows,
        "created_at": job.created_at,
        "completed_at": job.completed_at,
        "error_details": json.loads(job.error_details) if job.error_details else [],
    }


@router.get("/import-jobs/{job_id}/errors", summary="Tải danh sách lỗi import")
async def get_import_errors(job_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(404, "Import job không tồn tại")

    errors = json.loads(job.error_details) if job.error_details else []
    return {"job_id": job_id, "total_errors": len(errors), "errors": errors}


# ════════════════════════════════════════════════════════════════════════════
# QUERY PHƯƠNG TIỆN — Filter linh hoạt
# ════════════════════════════════════════════════════════════════════════════

def _build_pt_query(q, bien_so, hang_xe, loai_hinh, trang_thai, so_cho, so_cho_min, so_cho_max,
                    loai_so_huu=None, loai_di_thue=None,
                    han_dang_kiem_truoc=None, han_bao_hiem_truoc=None):
    stmt = select(PhuongTien)
    if q:
        q_like = f"%{q}%"
        stmt = stmt.where(or_(
            PhuongTien.bien_so.ilike(q_like),
            PhuongTien.hang_xe.ilike(q_like),
            PhuongTien.loai_hinh_hoat_dong.ilike(q_like),
            PhuongTien.tuyen_khai_thac.ilike(q_like),
        ))
    if bien_so:     stmt = stmt.where(PhuongTien.bien_so.ilike(f"%{bien_so}%"))
    if hang_xe:     stmt = stmt.where(PhuongTien.hang_xe.ilike(f"%{hang_xe}%"))
    if loai_hinh:   stmt = stmt.where(PhuongTien.loai_hinh_hoat_dong.ilike(f"%{loai_hinh}%"))
    if trang_thai:  stmt = stmt.where(PhuongTien.trang_thai == trang_thai)
    if so_cho:      stmt = stmt.where(PhuongTien.so_cho == so_cho)
    if so_cho_min:  stmt = stmt.where(PhuongTien.so_cho >= so_cho_min)
    if so_cho_max:  stmt = stmt.where(PhuongTien.so_cho <= so_cho_max)
    if loai_so_huu == "X":  stmt = stmt.where(PhuongTien.loai_so_huu != None)
    if loai_so_huu == "khong":  stmt = stmt.where(PhuongTien.loai_so_huu == None)
    if loai_di_thue == "X":  stmt = stmt.where(PhuongTien.loai_di_thue != None)
    if loai_di_thue == "khong":  stmt = stmt.where(PhuongTien.loai_di_thue == None)
    if han_dang_kiem_truoc:  stmt = stmt.where(PhuongTien.han_dang_kiem <= han_dang_kiem_truoc)
    if han_bao_hiem_truoc:   stmt = stmt.where(PhuongTien.han_bao_hiem <= han_bao_hiem_truoc)
    return stmt


def _build_lx_query(q, ho_ten, hang_gplx, trang_thai, gplx_het_han_truoc,
                    nhiem_vu=None, dong_bhxh_bhyt=None, ksk_ket_qua=None, ksk_het_han_truoc=None):
    stmt = select(LaiXe)
    if q:
        q_like = f"%{q}%"
        stmt = stmt.where(or_(
            LaiXe.ho_ten.ilike(q_like),
            LaiXe.hang_gplx.ilike(q_like),
            LaiXe.tap_huan_don_vi.ilike(q_like),
        ))
    if ho_ten:                stmt = stmt.where(LaiXe.ho_ten.ilike(f"%{ho_ten}%"))
    if hang_gplx:             stmt = stmt.where(LaiXe.hang_gplx == hang_gplx)
    if trang_thai:            stmt = stmt.where(LaiXe.trang_thai == trang_thai)
    if gplx_het_han_truoc:    stmt = stmt.where(LaiXe.han_gplx <= gplx_het_han_truoc)
    if nhiem_vu == "lai_xe":  stmt = stmt.where(LaiXe.nhiem_vu_lai_xe != None)
    if nhiem_vu == "nv_phuc_vu": stmt = stmt.where(LaiXe.nhiem_vu_nv_phuc_vu != None)
    if dong_bhxh_bhyt:        stmt = stmt.where(LaiXe.dong_bhxh_bhyt.ilike(f"%{dong_bhxh_bhyt}%"))
    if ksk_ket_qua:           stmt = stmt.where(LaiXe.ksk_ket_qua.ilike(f"%{ksk_ket_qua}%"))
    if ksk_het_han_truoc:     stmt = stmt.where(LaiXe.ksk_ngay_kham <= ksk_het_han_truoc)
    return stmt

@router.get("/phuong-tien", summary="Danh sách phương tiện với filter linh hoạt")
async def list_phuong_tien(
    q:            Optional[str] = Query(None),
    bien_so:      Optional[str] = Query(None),
    hang_xe:      Optional[str] = Query(None),
    loai_hinh:    Optional[str] = Query(None),
    trang_thai:   Optional[str] = Query(None),
    so_cho:       Optional[int] = Query(None),
    so_cho_min:   Optional[int] = Query(None),
    so_cho_max:   Optional[int] = Query(None),
    loai_so_huu:  Optional[str] = Query(None),
    loai_di_thue: Optional[str] = Query(None),
    han_dang_kiem_truoc: Optional[str] = Query(None),
    han_bao_hiem_truoc:  Optional[str] = Query(None),
    page:    int = Query(default=1, ge=1),
    size:    int = Query(default=20, ge=1, le=500),
    sort_by: str = Query(default="bien_so"),
    order:   str = Query(default="asc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
):
    stmt = _build_pt_query(q, bien_so, hang_xe, loai_hinh, trang_thai, so_cho, so_cho_min, so_cho_max,
                           loai_so_huu, loai_di_thue, han_dang_kiem_truoc, han_bao_hiem_truoc)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar()

    sort_col = getattr(PhuongTien, sort_by, PhuongTien.bien_so)
    stmt = stmt.order_by(sort_col.asc() if order == "asc" else sort_col.desc())
    stmt = stmt.offset((page - 1) * size).limit(size)
    rows = (await db.execute(stmt)).scalars().all()

    return {
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size,
        "data": [_pt_to_dict(r) for r in rows],
    }


@router.get("/phuong-tien/export", summary="Xuất Excel danh sách phương tiện (theo filter)")
async def export_phuong_tien(
    q:            Optional[str] = Query(None),
    bien_so:      Optional[str] = Query(None),
    hang_xe:      Optional[str] = Query(None),
    loai_hinh:    Optional[str] = Query(None),
    trang_thai:   Optional[str] = Query(None),
    so_cho:       Optional[int] = Query(None),
    so_cho_min:   Optional[int] = Query(None),
    so_cho_max:   Optional[int] = Query(None),
    loai_so_huu:  Optional[str] = Query(None),
    loai_di_thue: Optional[str] = Query(None),
    han_dang_kiem_truoc: Optional[str] = Query(None),
    han_bao_hiem_truoc:  Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = _build_pt_query(q, bien_so, hang_xe, loai_hinh, trang_thai, so_cho, so_cho_min, so_cho_max,
                           loai_so_huu, loai_di_thue, han_dang_kiem_truoc, han_bao_hiem_truoc)
    stmt = stmt.order_by(PhuongTien.bien_so)
    rows = (await db.execute(stmt)).scalars().all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phương tiện"

    headers = [
        "Biển số", "Hãng xe", "Năm SX", "Số chỗ", "Màu xe",
        "Loại hình HĐ", "Tuyến khai thác",
        "Hạn đăng kiểm", "Hạn phù hiệu", "Hạn BH TNDS",
        "GSHT tên", "GSHT đơn vị",
        "Sở hữu", "Đi thuê", "Trạng thái", "Ghi chú",
    ]
    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for ri, r in enumerate(rows, 2):
        ws.cell(ri, 1, r.bien_so)
        ws.cell(ri, 2, r.hang_xe)
        ws.cell(ri, 3, r.nam_san_xuat)
        ws.cell(ri, 4, r.so_cho)
        ws.cell(ri, 5, r.mau_xe)
        ws.cell(ri, 6, r.loai_hinh_hoat_dong)
        ws.cell(ri, 7, r.tuyen_khai_thac)
        ws.cell(ri, 8, r.han_dang_kiem)
        ws.cell(ri, 9, r.han_phu_hieu)
        ws.cell(ri, 10, r.han_bao_hiem)
        ws.cell(ri, 11, r.gsht_ten)
        ws.cell(ri, 12, r.gsht_don_vi)
        ws.cell(ri, 13, r.loai_so_huu)
        ws.cell(ri, 14, r.loai_di_thue)
        ws.cell(ri, 15, r.trang_thai)
        ws.cell(ri, 16, r.ghi_chu)

    col_widths = [14, 18, 8, 8, 10, 22, 26, 14, 14, 14, 20, 20, 10, 10, 14, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    fname = f"phuong_tien_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/phuong-tien/{pt_id}", summary="Chi tiết phương tiện")
async def get_phuong_tien(pt_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(PhuongTien).where(PhuongTien.id == pt_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Phương tiện không tồn tại")
    return _pt_to_dict(row)


# ════════════════════════════════════════════════════════════════════════════
# QUERY LÁI XE — Filter linh hoạt
# ════════════════════════════════════════════════════════════════════════════

@router.get("/lai-xe", summary="Danh sách lái xe với filter linh hoạt")
async def list_lai_xe(
    q:              Optional[str] = Query(None),
    ho_ten:         Optional[str] = Query(None),
    hang_gplx:      Optional[str] = Query(None),
    trang_thai:     Optional[str] = Query(None),
    gplx_het_han_truoc: Optional[str] = Query(None),
    nhiem_vu:       Optional[str] = Query(None, description="lai_xe | nv_phuc_vu"),
    dong_bhxh_bhyt: Optional[str] = Query(None, description="Có | Không"),
    ksk_ket_qua:    Optional[str] = Query(None, description="Đủ sức khỏe | Không đủ"),
    ksk_het_han_truoc: Optional[str] = Query(None),
    page:  int = Query(default=1, ge=1),
    size:  int = Query(default=20, ge=1, le=500),
    sort_by: str = Query(default="ho_ten"),
    order:   str = Query(default="asc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
):
    stmt = _build_lx_query(q, ho_ten, hang_gplx, trang_thai, gplx_het_han_truoc,
                           nhiem_vu, dong_bhxh_bhyt, ksk_ket_qua, ksk_het_han_truoc)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar()

    sort_col = getattr(LaiXe, sort_by, LaiXe.ho_ten)
    stmt = stmt.order_by(sort_col.asc() if order == "asc" else sort_col.desc())
    stmt = stmt.offset((page - 1) * size).limit(size)
    rows = (await db.execute(stmt)).scalars().all()

    return {
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size,
        "data": [_lx_to_dict(r) for r in rows],
    }


@router.get("/lai-xe/export", summary="Xuất Excel danh sách lái xe (theo filter)")
async def export_lai_xe(
    q:              Optional[str] = Query(None),
    ho_ten:         Optional[str] = Query(None),
    hang_gplx:      Optional[str] = Query(None),
    trang_thai:     Optional[str] = Query(None),
    gplx_het_han_truoc: Optional[str] = Query(None),
    nhiem_vu:       Optional[str] = Query(None),
    dong_bhxh_bhyt: Optional[str] = Query(None),
    ksk_ket_qua:    Optional[str] = Query(None),
    ksk_het_han_truoc: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    stmt = _build_lx_query(q, ho_ten, hang_gplx, trang_thai, gplx_het_han_truoc,
                           nhiem_vu, dong_bhxh_bhyt, ksk_ket_qua, ksk_het_han_truoc)
    stmt = stmt.order_by(LaiXe.ho_ten)
    rows = (await db.execute(stmt)).scalars().all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lái xe"

    headers = [
        "Họ tên", "Nhiệm vụ lái xe", "NV phục vụ",
        "Hạng GPLX", "Hạn GPLX",
        "Ngày ký HĐ", "Loại HĐ", "BHXH/BHYT",
        "Ngày khám SK", "Kết quả SK",
        "Ngày tập huấn", "Đơn vị TH", "Số GCN TH",
        "Trạng thái", "Ghi chú",
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ri, r in enumerate(rows, 2):
        ws.cell(ri, 1, r.ho_ten)
        ws.cell(ri, 2, r.nhiem_vu_lai_xe or "")
        ws.cell(ri, 3, r.nhiem_vu_nv_phuc_vu or "")
        ws.cell(ri, 4, r.hang_gplx)
        ws.cell(ri, 5, r.han_gplx)
        ws.cell(ri, 6, r.hop_dong_ngay_ky)
        ws.cell(ri, 7, r.hop_dong_loai)
        ws.cell(ri, 8, r.dong_bhxh_bhyt)
        ws.cell(ri, 9, r.ksk_ngay_kham)
        ws.cell(ri, 10, r.ksk_ket_qua)
        ws.cell(ri, 11, r.tap_huan_ngay)
        ws.cell(ri, 12, r.tap_huan_don_vi)
        ws.cell(ri, 13, r.tap_huan_so_gcn)
        ws.cell(ri, 14, r.trang_thai)
        ws.cell(ri, 15, r.ghi_chu)

    col_widths = [24, 14, 14, 12, 14, 14, 22, 12, 14, 18, 14, 22, 16, 14, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    fname = f"lai_xe_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/lai-xe/{lx_id}", summary="Chi tiết lái xe")
async def get_lai_xe(lx_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(LaiXe).where(LaiXe.id == lx_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(404, "Lái xe không tồn tại")
    return _lx_to_dict(row)


# ════════════════════════════════════════════════════════════════════════════
# STATS
# ════════════════════════════════════════════════════════════════════════════

@router.get("/stats", summary="Thống kê tổng hợp phương tiện & lái xe")
async def get_fleet_stats(db: AsyncSession = Depends(get_db)):
    # Phương tiện
    total_pt = (await db.execute(select(func.count(PhuongTien.id)))).scalar()
    pt_by_so_cho = (await db.execute(
        select(PhuongTien.so_cho, func.count().label("count"))
        .group_by(PhuongTien.so_cho)
        .order_by(PhuongTien.so_cho)
    )).all()
    pt_by_trang_thai = (await db.execute(
        select(PhuongTien.trang_thai, func.count().label("count"))
        .group_by(PhuongTien.trang_thai)
    )).all()

    # Lái xe
    total_lx = (await db.execute(select(func.count(LaiXe.id)))).scalar()
    lx_by_hang_gplx = (await db.execute(
        select(LaiXe.hang_gplx, func.count().label("count"))
        .group_by(LaiXe.hang_gplx)
        .order_by(LaiXe.hang_gplx)
    )).all()
    lx_by_trang_thai = (await db.execute(
        select(LaiXe.trang_thai, func.count().label("count"))
        .group_by(LaiXe.trang_thai)
    )).all()

    return {
        "phuong_tien": {
            "total": total_pt,
            "by_so_cho": [{"so_cho": r[0], "count": r[1]} for r in pt_by_so_cho],
            "by_trang_thai": [{"trang_thai": r[0], "count": r[1]} for r in pt_by_trang_thai],
        },
        "lai_xe": {
            "total": total_lx,
            "by_hang_gplx": [{"hang": r[0], "count": r[1]} for r in lx_by_hang_gplx],
            "by_trang_thai": [{"trang_thai": r[0], "count": r[1]} for r in lx_by_trang_thai],
        },
    }


# ════════════════════════════════════════════════════════════════════════════
# BACKGROUND IMPORT TASK
# ════════════════════════════════════════════════════════════════════════════

async def _run_import(job_id: str, filepath: str, table: str, header_row: int, data_start_row: int, preview: bool = False):
    from app.core.database import AsyncSessionLocal

    stats = {"total": 0, "success": 0, "errors": 0, "error_details": []}
    async with AsyncSessionLocal() as db:
        try:
            # Update status → processing
            result = await db.execute(select(ImportJob).where(ImportJob.id == job_id))
            job = result.scalar_one()
            job.status = ImportStatus.PROCESSING
            await db.commit()

            # Run import
            stats = await import_excel(db, filepath, table, job_id, header_row, data_start_row, preview)
            await db.commit()  # commit data import vào DB

            # Update job kết quả
            async with AsyncSessionLocal() as db2:
                result2 = await db2.execute(select(ImportJob).where(ImportJob.id == job_id))
                job2 = result2.scalar_one()
                job2.status = ImportStatus.DONE
                job2.total_rows = stats["total"]
                job2.success_rows = stats["success"]
                job2.error_rows = stats["errors"]
                job2.error_details = json.dumps(stats["error_details"], ensure_ascii=False) if stats["error_details"] else None
                job2.completed_at = datetime.now(timezone.utc)
                await db2.commit()

        except Exception as e:
            async with AsyncSessionLocal() as db_err:
                result3 = await db_err.execute(select(ImportJob).where(ImportJob.id == job_id))
                job3 = result3.scalar_one_or_none()
                if job3:
                    job3.status = ImportStatus.FAILED
                    job3.error_details = json.dumps([{"error": str(e)}])
                    job3.completed_at = datetime.now(timezone.utc)
                    await db_err.commit()

    return stats


# ─── Helpers ────────────────────────────────────────────────────────────────
def _pt_to_dict(r: PhuongTien) -> dict:
    return {
        "id": r.id,
        "bien_so": r.bien_so,
        "hang_xe": r.hang_xe,
        "nam_san_xuat": r.nam_san_xuat,
        "so_cho": r.so_cho,
        "mau_xe": r.mau_xe,
        "loai_hinh_hoat_dong": r.loai_hinh_hoat_dong,
        "tuyen_khai_thac": r.tuyen_khai_thac,
        "han_dang_kiem": r.han_dang_kiem,
        "han_phu_hieu": r.han_phu_hieu,
        "han_bao_hiem": r.han_bao_hiem,
        "gsht_ten": r.gsht_ten,
        "gsht_don_vi": r.gsht_don_vi,
        "loai_so_huu": r.loai_so_huu,
        "loai_di_thue": r.loai_di_thue,
        "trang_thai": r.trang_thai,
        "ghi_chu": r.ghi_chu,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }

def _lx_to_dict(r: LaiXe) -> dict:
    return {
        "id": r.id,
        "ho_ten": r.ho_ten,
        "nhiem_vu_lai_xe": r.nhiem_vu_lai_xe,
        "nhiem_vu_nv_phuc_vu": r.nhiem_vu_nv_phuc_vu,
        "hang_gplx": r.hang_gplx,
        "han_gplx": r.han_gplx,
        "hop_dong_ngay_ky": r.hop_dong_ngay_ky,
        "hop_dong_loai": r.hop_dong_loai,
        "dong_bhxh_bhyt": r.dong_bhxh_bhyt,
        "ksk_ngay_kham": r.ksk_ngay_kham,
        "ksk_ket_qua": r.ksk_ket_qua,
        "tap_huan_ngay": r.tap_huan_ngay,
        "tap_huan_don_vi": r.tap_huan_don_vi,
        "tap_huan_so_gcn": r.tap_huan_so_gcn,
        "trang_thai": r.trang_thai,
        "ghi_chu": r.ghi_chu,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }
